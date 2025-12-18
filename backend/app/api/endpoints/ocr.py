from fastapi import APIRouter, UploadFile, File, HTTPException, Body
from fastapi.responses import StreamingResponse
from app.services.ocr_engine import ocr_engine
from app.services.llm_engine import llm_engine
import pandas as pd
from io import BytesIO
from urllib.parse import quote
from openpyxl.styles import PatternFill, Font
import time
import re

router = APIRouter()

# ==========================================
# 1. 核心分析接口
# ==========================================
@router.post("/analyze")
async def analyze_document(file: UploadFile = File(...)):
    if not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="File must be an image")
    
    print(f"\n====== 📸 收到新图片: {file.filename} ======")
    start_time = time.time()
    
    try:
        print("   1. 正在读取文件...")
        content = await file.read()
        
        print("   2. 🚀 启动 PaddleOCR...")
        raw_text = ocr_engine.extract_text(content)
        
        if not raw_text:
            return {"status": "failed", "message": "未识别到任何文字"}
            
        print("   3. 🧠 发送给火山引擎 (v6.0) 进行清洗...")
        cleaned_data = llm_engine.parse_content(raw_text)
        
        total_cost = time.time() - start_time
        print(f"====== 🎉 处理结束! 总耗时: {total_cost:.2f}秒 ======\n")
        
        return {
            "status": "success",
            "filename": file.filename,
            "data": cleaned_data.get("data", {}),
            "raw_text": raw_text
        }
    except Exception as e:
        print(f"❌ 处理出错: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ==========================================
# 2. 导出 Excel 接口
# ==========================================
@router.post("/export-excel")
async def export_to_excel(data: dict = Body(...)):
    print("\n====== 📊 收到导出 Excel 请求 ======")
    try:
        store_name = data.get("data", {}).get("store_name", "库存盘点单")
        items = data.get("data", {}).get("items", [])
        
        if not items and isinstance(data.get("data"), list):
             items = data.get("data")

        if not items:
            return {"status": "error", "message": "没有数据可导出"}

        print(f"   1. 正在处理 {len(items)} 条数据...")

        # --- 1. 动态生成 DataFrame ---
        df = pd.DataFrame(items)
        
        # 全量映射表
        column_mapping = {
            "code": "商品编码",
            "batch_number": "批次",
            "name": "商品名称",
            "spec": "规格",
            "unit": "单位",
            "system_stock": "系统库存",
            "actual_count": "实盘数量"
        }

        # 动态列筛选
        existing_columns = [col for col in column_mapping.keys() if col in df.columns]
        df = df[existing_columns] 
        df = df.rename(columns=column_mapping)

        # --- 2. 写入 Excel ---
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            sheet_name = "盘点数据"
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            
            # --- 3. 智能上色 ---
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            red_font = Font(color="FF0000", bold=True)
            green_font = Font(color="008000", bold=True)

            header_row = [cell.value for cell in worksheet[1]]
            
            actual_idx = header_row.index("实盘数量") + 1 if "实盘数量" in header_row else -1
            system_idx = header_row.index("系统库存") + 1 if "系统库存" in header_row else -1

            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                actual_cell = row[actual_idx - 1] if actual_idx > 0 else None
                system_cell = row[system_idx - 1] if system_idx > 0 else None
                
                actual_val = actual_cell.value if actual_cell else None
                system_val = system_cell.value if system_cell else None

                # 规则 1: 漏填 (标黄)
                if actual_cell and (actual_val is None or actual_val == ""):
                    for cell in row:
                        cell.fill = yellow_fill
                    actual_cell.value = "待核查"
                    actual_cell.font = Font(color="FF9900", italic=True)

                # 规则 2: 差异 (标红)
                elif system_cell and actual_cell:
                    try:
                        sys_num = float(re.findall(r"\d+\.?\d*", str(system_val))[0])
                        act_num = float(re.findall(r"\d+\.?\d*", str(actual_val))[0])
                        if sys_num != act_num:
                            actual_cell.font = red_font
                        else:
                            actual_cell.font = green_font
                    except:
                        pass

            for i, col in enumerate(df.columns):
                col_letter = chr(65 + i)
                if "名称" in col: width = 30
                elif "编码" in col: width = 20
                elif "单位" in col: width = 8
                else: width = 15
                worksheet.column_dimensions[col_letter].width = width

        output.seek(0)
        filename = f"{store_name}_导出.xlsx"
        filename_encoded = quote(filename)
        
        print(f"====== ✅ Excel 生成完毕: {filename} ======\n")

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=utf-8''{filename_encoded}"
            }
        )

    except Exception as e:
        print(f"❌ 导出 Excel 失败: {e}")
        return {"status": "error", "message": str(e)}